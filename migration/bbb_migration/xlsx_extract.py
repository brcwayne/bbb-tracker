"""xlsm ham satır çıkarımı — okuma dışında iş kuralı yok."""
from __future__ import annotations

import openpyxl

from .constants import DATA_START_ROW


def _load(path):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _cell(ws, col, row):
    return ws[f"{col}{row}"].value


def _first_nonempty(*vals):
    for v in vals:
        if v is not None and not (isinstance(v, str) and not v.strip()):
            return v
    return None


def extract_trades(path) -> list[dict]:
    wb = _load(path)
    ws = wb["Trade Log"]
    out = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        code = _cell(ws, "F", row)
        if code is None or (isinstance(code, str) and not code.strip()):
            continue
        out.append({
            "row_no": row,
            "portfoy_raw": _cell(ws, "D", row),
            "tarih_raw": _cell(ws, "E", row),
            "kod_raw": code,
            "yon_raw": _cell(ws, "G", row),
            "tl_raw": _cell(ws, "H", row),
            "fiyat_raw": _cell(ws, "I", row),
            "lot_raw": _cell(ws, "J", row),
            "komisyon_raw": _cell(ws, "K", row),
        })
    wb.close()
    return out


def extract_bank_transfers(path) -> list[dict]:
    wb = _load(path)
    ws = wb["Bank Transfers"]
    out = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        action = _cell(ws, "E", row)
        date = _cell(ws, "D", row)
        if action is None and date is None:
            continue
        out.append({
            "row_no": row,
            "tarih_raw": date,
            "action_raw": action,
            "gross_raw": _cell(ws, "F", row),
            "fees_raw": _cell(ws, "G", row),
            "net_raw": _cell(ws, "H", row),
            "notes_raw": _cell(ws, "I", row),
        })
    wb.close()
    return out


def extract_dividends(path) -> list[dict]:
    wb = _load(path)
    ws = wb["Dividends"]
    out = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        code = _cell(ws, "D", row)
        if code is None or (isinstance(code, str) and not code.strip()):
            continue
        # Tarih: ilk dolu olan -> H (Ex-Div, bu workbook'ta hep boş) / J (Payable) / I (Record)
        resolved_date = _first_nonempty(
            _cell(ws, "H", row), _cell(ws, "J", row), _cell(ws, "I", row)
        )
        out.append({
            "row_no": row,
            "kod_raw": code,
            "tur_raw": _cell(ws, "E", row),
            "value_raw": _cell(ws, "F", row),
            "usdtry_raw": _cell(ws, "G", row),
            "exdiv_raw": resolved_date,
            "paid_usd_raw": _cell(ws, "K", row),
        })
    wb.close()
    return out


def extract_reference(path) -> dict:
    wb = _load(path)
    out = {"stock_position": [], "monthly_report": [], "portfoyler": []}
    if "Bank Transfers" in wb.sheetnames:
        out["total_deposits"] = wb["Bank Transfers"]["H5"].value
    if "Stock Position" in wb.sheetnames:
        ws = wb["Stock Position"]
        for row in range(DATA_START_ROW, ws.max_row + 1):
            code = _cell(ws, "D", row)
            if not isinstance(code, str) or not code.strip():
                continue
            out["stock_position"].append({
                "kod": code.strip(),
                "lot": _cell(ws, "E", row),
                "ave_price": _cell(ws, "F", row),
                "amount": _cell(ws, "G", row),
            })
    if "Monthly Report" in wb.sheetnames:
        ws = wb["Monthly Report"]
        for row in range(21, ws.max_row + 1):
            m = _cell(ws, "D", row)
            if not hasattr(m, "year") or _cell(ws, "C", row) in (None, ""):
                continue
            out["monthly_report"].append({
                "ay": f"{m.year:04d}-{m.month:02d}",
                "beg_capital": _cell(ws, "E", row),
                "deposits": _cell(ws, "F", row),
                "withdrawals": _cell(ws, "H", row),
                "gain": _cell(ws, "I", row),
                "loss": _cell(ws, "J", row),
                "cash_div": _cell(ws, "K", row),
                "end_capital": _cell(ws, "L", row),
                "tax_fees": _cell(ws, "M", row),
            })
    wb.close()
    return out
