"""Deterministik test workbook'u. Çalıştır: python tests/fixtures/build_fixture.py"""
import datetime as dt
from pathlib import Path
import openpyxl

OUT = Path(__file__).with_name("mini.xlsm")


def _set(ws, cell, val):
    ws[cell] = val


def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tl = wb.create_sheet("Trade Log")
    _set(tl, "C14", "No."); _set(tl, "E14", "Date"); _set(tl, "F14", "Stock Code")
    _set(tl, "G14", "Action"); _set(tl, "H14", "TL"); _set(tl, "I14", "Price")
    _set(tl, "J14", "Total Shares"); _set(tl, "K14", "Manual Fees")
    rows = [
        # row, portfoy, date, code, action, tl, price_usd, lot, fee
        (15, "M.Alfa", dt.datetime(2020, 1, 6), "ASTOR", "BUY", None, 1.00, 100, 0),
        (16, "M.Alfa", dt.datetime(2020, 6, 10), "ASTOR", "BUY", None, 2.00, 100, 0),
        (17, "M.Alfa", dt.datetime(2021, 3, 15), "ASTOR", "SELL", None, 5.00, 50, 0),
        (18, "KASA", dt.datetime(2019, 7, 1), "XAU", "BUY", None, 50.00, 10, 0),
        (19, "QNB.F", dt.datetime(2022, 2, 3), "AFT.F", "BUY", None, 1.00, 1000, 0),
        (20, "GARAN", dt.datetime(2023, 11, 20), "THYAO", "BUY", 1200.00, 40.00, 25, 1.5),
        (21, "M.Alfa", dt.datetime(2021, 3, 15), "ASTOR", "SELL", None, 6.00, 200, 0),  # aşırı satış → recon hata
    ]
    for r, pf, d, code, act, tl_, px, lot, fee in rows:
        _set(tl, f"A{r}", "#VALUE!")
        _set(tl, f"C{r}", r - 14); _set(tl, f"D{r}", pf); _set(tl, f"E{r}", d)
        _set(tl, f"F{r}", code); _set(tl, f"G{r}", act)
        if tl_ is not None:
            _set(tl, f"H{r}", tl_)
        _set(tl, f"I{r}", px); _set(tl, f"J{r}", lot); _set(tl, f"K{r}", fee)
    # gürültü satırı: sadece yardımcı sütun dolu
    _set(tl, "A25", "#VALUE!")

    # Bank Transfers — gerçek workbook düzeni: veri C sütunundan başlar.
    bt = wb.create_sheet("Bank Transfers")
    _set(bt, "C14", "No."); _set(bt, "D14", "Date"); _set(bt, "E14", "Action")
    _set(bt, "F14", "Gross Amount"); _set(bt, "G14", "Fees"); _set(bt, "H14", "Net Amount"); _set(bt, "I14", "Notes")
    bt_rows = [
        (15, dt.datetime(2019, 1, 2), "Deposit", 1000.00, 0, 1000.00, "ilk"),
        (16, dt.datetime(2021, 5, 5), "Withdraw", 200.00, 0, 200.00, "cekim"),
    ]
    for r, d, act, g, f, net, note in bt_rows:
        _set(bt, f"C{r}", r - 14); _set(bt, f"D{r}", d); _set(bt, f"E{r}", act)
        _set(bt, f"F{r}", g); _set(bt, f"G{r}", f); _set(bt, f"H{r}", net); _set(bt, f"I{r}", note)
    # gürültü: sadece index sütunu dolu (gerçek workbook'taki boş şablon satırları)
    _set(bt, "C17", 3)

    # Dividends — gerçek workbook düzeni: veri C'den başlar, tarih J (Payable), paid USD K.
    # H (Ex-Div) bu workbook'ta hep boş → extractor J'ye düşmeli.
    dv = wb.create_sheet("Dividends")
    _set(dv, "C14", "No."); _set(dv, "D14", "Stock Code"); _set(dv, "E14", "Type")
    _set(dv, "F14", "Value / Description"); _set(dv, "G14", "usdtry"); _set(dv, "H14", "Ex-Div Date")
    _set(dv, "I14", "Record date"); _set(dv, "J14", "Payable date"); _set(dv, "K14", "Paid Value")
    _set(dv, "C15", 1); _set(dv, "D15", "ASTOR"); _set(dv, "E15", "Cash")
    _set(dv, "F15", 100.00); _set(dv, "G15", 25.00)
    # H15 (Ex-Div) kasıtlı boş; tarih J15'ten gelmeli
    _set(dv, "J15", dt.datetime(2023, 4, 10)); _set(dv, "K15", 4.00)
    # gürültü: sadece index sütunu dolu boş şablon satırı
    _set(dv, "C16", 2)

    # Stock Position — gerçek workbook düzeni: veri C'den başlar (C=No, D=Code, E=Shares, F=Ave, G=Amount).
    sp = wb.create_sheet("Stock Position")
    _set(sp, "C14", "No."); _set(sp, "D14", "Stock Code"); _set(sp, "E14", "Total\nShares")
    _set(sp, "F14", "Ave.\nPrice"); _set(sp, "G14", "Amount"); _set(sp, "H14", "Weight %")
    # ASTOR migration sonrası: 100+100-50-200 = -150 (fixture kasıtlı hatalı); referans "doğru" değeri 150 lot der
    sp_rows = [
        (15, "ASTOR", 150, 1.5, 225.0),
        (16, "XAU", 10, 50.0, 500.0),
        (17, "AFT.F", 1000, 1.0, 1000.0),
        (18, "THYAO", 25, 40.06, 1001.5),
    ]
    for r, code, lot, ave, amt in sp_rows:
        _set(sp, f"C{r}", r - 14); _set(sp, f"D{r}", code); _set(sp, f"E{r}", lot)
        _set(sp, f"F{r}", ave); _set(sp, f"G{r}", amt)

    # Monthly Report — gerçek workbook düzeni: D=Month, E=Beg, F=Net Dep/Withdr, G=Additional,
    # H=Withdrawals, I=Gain, J=Loss, K=Cash Div, L=END CAPITAL, M=TAX & FEES.
    # Satır 21 = YTD TOTALS (C boş), veri satırları 22+.
    mr = wb.create_sheet("Monthly Report")
    _set(mr, "D20", "Month"); _set(mr, "E20", "BEGINNING CAPITAL")
    _set(mr, "F20", "Net Deposits/ Withdrawals"); _set(mr, "G20", "Additional Deposits")
    _set(mr, "H20", "Withdrawals"); _set(mr, "I20", "Gain"); _set(mr, "J20", "Loss")
    _set(mr, "K20", "Cash Dividends"); _set(mr, "L20", "END CAPITAL"); _set(mr, "M20", "TAX & FEES")
    # row, month, beg, net_dep, additional, withdrawals, gain, loss, cash_div, end_capital, tax_fees
    mr_rows = [
        (21, dt.datetime(2022, 6, 1), 1000.0, 0.0, 0.0, 0.0, 175.0, 0.0, 0.0, 1075.0, 0.0),  # TOTALS (C boş)
        (22, dt.datetime(2020, 1, 1), 1000.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 900.0, 0.0),
        (23, dt.datetime(2021, 3, 1), 900.0, 0.0, 0.0, 0.0, 175.0, 0.0, 0.0, 1075.0, 0.0),
    ]
    for r, m, beg, nd, addl, wd, gn, ls, cd, endc, tf in mr_rows:
        if r != 21:
            _set(mr, f"C{r}", r - 21)  # veri satırlarında index; TOTALS satırında C boş
        _set(mr, f"D{r}", m); _set(mr, f"E{r}", beg); _set(mr, f"F{r}", nd); _set(mr, f"G{r}", addl)
        _set(mr, f"H{r}", wd); _set(mr, f"I{r}", gn); _set(mr, f"J{r}", ls)
        _set(mr, f"K{r}", cd); _set(mr, f"L{r}", endc); _set(mr, f"M{r}", tf)

    wb.save(OUT)
    print("wrote", OUT)


if __name__ == "__main__":
    build()
